#!/usr/bin/env python3
"""
Extract active employees from the FY2026 Staff Costs Excel workbook
and write them as a JSON fixture file.

Input:  data/budgets/02_EFIR_Staff_Costs_FY2026.xlsx  (Master Data sheet)
Output: data/fixtures/fy2026-staff-costs.json

Usage:  python3 scripts/extract-staff-costs.py
"""

import json
import os
import sys
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(ROOT, 'data', 'budgets', '02_EFIR_Staff_Costs_FY2026.xlsx')
EXISTING_FIXTURE_PATH = os.path.join(ROOT, 'data', 'fixtures', 'fy2026-staff-costs.json')
OUTPUT_PATH = os.path.join(ROOT, 'data', 'fixtures', 'fy2026-staff-costs.json')

# Column indices in the Master Data sheet (1-based)
COL_NUM = 1
COL_LAST_NAME = 2
COL_FIRST_NAME = 3
COL_CONTRACT_TYPE = 4
COL_CONTRACT_STATUS = 5
COL_FUNCTION_ROLE = 6
COL_DEPARTMENT = 7
COL_SUBJECT = 8
COL_TEACHER_TYPE = 9
COL_TEACHER_PROFILE = 10
COL_LEVEL = 11
COL_STATUS = 12
COL_NATIONALITY = 13
COL_JOINING_DATE = 14
COL_YOS = 15
COL_BASE_SALARY = 16
COL_HOUSING = 17
COL_TRANSPORT = 18
COL_RESP_PREMIUM = 19
COL_HSA = 20
COL_MONTHLY_GROSS = 21
COL_HOURLY_PCT = 22
COL_PAYMENT_METHOD = 23
COL_INCREASE = 24

HEADER_ROW = 4
DATA_START_ROW = 5

# Payment method mapping (French -> English)
PAYMENT_MAP = {
    'Virement': 'Bank Transfer',
    'Liquide': 'Cash',
    'Mudad': 'Mudad',
    'N/A': 'Bank Transfer',  # default for N/A
}

# Fallback homeBand derivation for employees not found in existing fixture.
# Uses department-based mapping per user specification, with level as secondary signal.
def derive_homeband(department, level):
    """Derive homeBand from Department (primary) and Level (secondary)."""
    dept_lower = department.lower() if department else ''
    level_lower = level.lower() if level else ''

    if 'maternelle' in dept_lower:
        return 'MATERNELLE'
    if 'élémentaire' in dept_lower or 'elementaire' in dept_lower:
        # Élémentaire dept employees may teach Maternelle levels
        if 'maternelle' in level_lower:
            return 'MATERNELLE'
        return 'ELEMENTAIRE'
    # Check collège before lycée because "Collège / Lycée" contains both,
    # and the majority of employees in that combined dept are COLLEGE-level.
    if 'collège' in dept_lower or 'college' in dept_lower:
        return 'COLLEGE'
    if 'lycée' in dept_lower or 'lycee' in dept_lower:
        return 'LYCEE'
    if 'vie scolaire' in dept_lower:
        return 'COLLEGE'
    if 'administration' in dept_lower:
        return None
    if 'direction' in dept_lower:
        return None
    # Anything else -> NON_ACADEMIC (per user spec: "others->NON_ACADEMIC")
    return 'NON_ACADEMIC'


def safe_decimal(value, decimals=4):
    """Convert a numeric value to a fixed-decimal string like '1234.0000'."""
    if value is None:
        return f'0.{"0" * decimals}'
    try:
        num = float(value)
    except (ValueError, TypeError):
        return f'0.{"0" * decimals}'
    return f'{num:.{decimals}f}'


def safe_str(value):
    """Convert a cell value to a stripped string, or empty string if None."""
    if value is None:
        return ''
    return str(value).strip()


def format_date(value):
    """Convert a datetime or string to 'YYYY-MM-DD' format."""
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        return value[:10]
    return '2024-09-01'  # fallback


def main():
    # Load existing fixture for homeBand carry-forward
    existing_homeband = {}
    existing_costmode = {}
    if os.path.exists(EXISTING_FIXTURE_PATH):
        with open(EXISTING_FIXTURE_PATH, 'r') as f:
            existing_data = json.load(f)
        for emp in existing_data:
            existing_homeband[emp['name']] = emp.get('homeBand')
            existing_costmode[emp['name']] = emp.get('costMode', 'LOCAL_PAYROLL')

    # Load Excel workbook
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Master Data']

    employees = []
    seq = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        status_raw = ws.cell(row=row_idx, column=COL_STATUS).value
        if status_raw is None:
            continue
        status_str = safe_str(status_raw)

        # Filter out Departed employees
        if status_str == 'Departed':
            continue

        seq += 1
        employee_code = f'EFIR-{seq:03d}'

        last_name = safe_str(ws.cell(row=row_idx, column=COL_LAST_NAME).value)
        first_name = safe_str(ws.cell(row=row_idx, column=COL_FIRST_NAME).value)
        name = f'{last_name} {first_name}'.strip()

        function_role = safe_str(ws.cell(row=row_idx, column=COL_FUNCTION_ROLE).value)
        department = safe_str(ws.cell(row=row_idx, column=COL_DEPARTMENT).value)
        subject = safe_str(ws.cell(row=row_idx, column=COL_SUBJECT).value)
        level = safe_str(ws.cell(row=row_idx, column=COL_LEVEL).value)
        nationality = safe_str(ws.cell(row=row_idx, column=COL_NATIONALITY).value)
        contract_type = safe_str(ws.cell(row=row_idx, column=COL_CONTRACT_TYPE).value)
        teacher_type = safe_str(ws.cell(row=row_idx, column=COL_TEACHER_TYPE).value)

        # costMode: Titulaire EN -> AEFE_RECHARGE, otherwise LOCAL_PAYROLL
        if contract_type == 'Titulaire EN':
            cost_mode = 'AEFE_RECHARGE'
        else:
            cost_mode = 'LOCAL_PAYROLL'

        # homeBand: carry forward from existing fixture if available,
        # otherwise derive from department/level for new employees
        if name in existing_homeband:
            home_band = existing_homeband[name]
        else:
            home_band = derive_homeband(department, level)

        # isTeaching: true if Teacher / Non-Teacher column says "Teacher" or "Enseignant"
        is_teaching = teacher_type.lower() in ('teacher', 'enseignant')

        # isSaudi
        is_saudi = nationality == 'Saudi'

        # isAjeer: true if non-Saudi (per user instructions)
        is_ajeer = not is_saudi

        # joiningDate
        joining_date_raw = ws.cell(row=row_idx, column=COL_JOINING_DATE).value
        joining_date = format_date(joining_date_raw)

        # paymentMethod
        payment_raw = safe_str(ws.cell(row=row_idx, column=COL_PAYMENT_METHOD).value)
        payment_method = PAYMENT_MAP.get(payment_raw, 'Bank Transfer')

        # hourlyPercentage
        hourly_raw = ws.cell(row=row_idx, column=COL_HOURLY_PCT).value
        hourly_pct = safe_decimal(hourly_raw if hourly_raw is not None else 1.0)

        # Salary fields
        base_salary_raw = ws.cell(row=row_idx, column=COL_BASE_SALARY).value
        base_salary = safe_decimal(base_salary_raw)

        housing_raw = ws.cell(row=row_idx, column=COL_HOUSING).value
        housing_allowance = safe_decimal(housing_raw)

        transport_raw = ws.cell(row=row_idx, column=COL_TRANSPORT).value
        transport_allowance = safe_decimal(transport_raw)

        resp_raw = ws.cell(row=row_idx, column=COL_RESP_PREMIUM).value
        responsibility_premium = safe_decimal(resp_raw)

        hsa_raw = ws.cell(row=row_idx, column=COL_HSA).value
        hsa_amount = safe_decimal(hsa_raw)

        # Augmentation: increase SAR / baseSalary -> percentage
        increase_raw = ws.cell(row=row_idx, column=COL_INCREASE).value
        increase_sar = 0.0
        try:
            increase_sar = float(increase_raw) if increase_raw else 0.0
        except (ValueError, TypeError):
            increase_sar = 0.0

        base_salary_num = float(base_salary_raw) if base_salary_raw else 0.0
        if increase_sar > 0 and base_salary_num > 0:
            augmentation = increase_sar / base_salary_num
            augmentation_str = f'{augmentation:.4f}'
            augmentation_effective_date = '2026-09-01'
        else:
            augmentation_str = '0.0000'
            augmentation_effective_date = None

        # Use "Existing" for all active employees (user instruction)
        status = 'Existing'

        # Level normalization: convert "Élémentaire" -> "Élémentaire" (keep as-is from Excel)
        # But the existing fixture uses accent-free "Elementaire" for level values
        # Let me check... the existing fixture has "Elementaire" not "Élémentaire"
        # Keep the raw level from Excel as-is

        employee = {
            'employeeCode': employee_code,
            'name': name,
            'functionRole': function_role,
            'department': department,
            'costMode': cost_mode,
            'subject': subject,
            'homeBand': home_band,
            'level': level if level else None,
            'status': status,
            'joiningDate': joining_date,
            'paymentMethod': payment_method,
            'isSaudi': is_saudi,
            'isAjeer': is_ajeer,
            'isTeaching': is_teaching,
            'hourlyPercentage': hourly_pct,
            'baseSalary': base_salary,
            'housingAllowance': housing_allowance,
            'transportAllowance': transport_allowance,
            'responsibilityPremium': responsibility_premium,
            'hsaAmount': hsa_amount,
            'augmentation': augmentation_str,
            'augmentationEffectiveDate': augmentation_effective_date,
            'recordType': 'EMPLOYEE',
        }

        employees.append(employee)

    # Sort by employeeCode
    employees.sort(key=lambda e: e['employeeCode'])

    # Write output
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(employees, f, indent='\t', ensure_ascii=False)
        f.write('\n')

    print(f'Wrote {len(employees)} active employees to {OUTPUT_PATH}')

    # Validation summary
    aefe_count = sum(1 for e in employees if e['costMode'] == 'AEFE_RECHARGE')
    local_count = sum(1 for e in employees if e['costMode'] == 'LOCAL_PAYROLL')
    saudi_count = sum(1 for e in employees if e['isSaudi'])
    ajeer_count = sum(1 for e in employees if e['isAjeer'])
    teaching_count = sum(1 for e in employees if e['isTeaching'])
    with_aug = sum(1 for e in employees if e['augmentation'] != '0.0000')
    null_band = sum(1 for e in employees if e['homeBand'] is None)
    print(f'  LOCAL_PAYROLL: {local_count}, AEFE_RECHARGE: {aefe_count}')
    print(f'  Saudi: {saudi_count}, Ajeer (non-Saudi): {ajeer_count}')
    print(f'  Teaching: {teaching_count}, Non-teaching: {len(employees) - teaching_count}')
    print(f'  With augmentation: {with_aug}')
    print(f'  Null homeBand: {null_band}')

    # Show employees that were NOT matched in existing fixture
    unmatched = [e for e in employees if e['name'] not in existing_homeband]
    if unmatched:
        print(f'\n  New employees (not in previous fixture): {len(unmatched)}')
        for e in unmatched:
            print(f'    {e["employeeCode"]} {e["name"]:35s} '
                  f'dept={e["department"]:30s} homeBand={e["homeBand"]!r}')


if __name__ == '__main__':
    main()
