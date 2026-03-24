#!/usr/bin/env python3
"""
Generate EFIR Subjects Coverage Excel file.
Cross-references AEFE/Éducation Nationale official curriculum with BudFin data.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
CYCLE_FILL = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2C3E50')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12, color='2C3E50')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='2C3E50')
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
CHECK_FONT = Font(name='Calibri', size=10, color='27AE60')
CROSS_FONT = Font(name='Calibri', size=10, color='E74C3C')
WARN_FONT = Font(name='Calibri', size=10, color='E67E22')

GREEN_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
RED_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

MATERNELLE_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
ELEM_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
COLLEGE_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
LYCEE_FILL = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='BDBDBD'),
    right=Side(style='thin', color='BDBDBD'),
    top=Side(style='thin', color='BDBDBD'),
    bottom=Side(style='thin', color='BDBDBD'),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header_row(ws, row, max_col, fill=None):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, font=None, fill=None, alignment=None):
    """Apply styling to a data cell."""
    cell = ws.cell(row=row, column=col)
    cell.font = font or NORMAL_FONT
    if fill:
        cell.fill = fill
    cell.alignment = alignment or LEFT_WRAP
    cell.border = THIN_BORDER
    return cell


def auto_width(ws, min_width=10, max_width=35):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                longest_line = max(len(line) for line in lines)
                max_len = max(max_len, longest_line)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ── Data ─────────────────────────────────────────────────────────────────────

# All levels
LEVELS = ['PS', 'MS', 'GS', 'CP', 'CE1', 'CE2', 'CM1', 'CM2',
          '6ème', '5ème', '4ème', '3ème', '2nde', '1ère', 'Terminale']

LEVEL_CYCLES = {
    'PS': 'Cycle 1 - Maternelle', 'MS': 'Cycle 1 - Maternelle',
    'GS': 'Cycle 1 - Maternelle',
    'CP': 'Cycle 2 - Apprentissages fondamentaux',
    'CE1': 'Cycle 2 - Apprentissages fondamentaux',
    'CE2': 'Cycle 2 - Apprentissages fondamentaux',
    'CM1': 'Cycle 3 - Consolidation', 'CM2': 'Cycle 3 - Consolidation',
    '6ème': 'Cycle 3 - Consolidation',
    '5ème': 'Cycle 4 - Approfondissements',
    '4ème': 'Cycle 4 - Approfondissements',
    '3ème': 'Cycle 4 - Approfondissements',
    '2nde': 'Lycée - Seconde (Détermination)',
    '1ère': 'Lycée - Première',
    'Terminale': 'Lycée - Terminale',
}

# Main matrix: subject → levels where taught
# Format: { official_name: { 'levels': {level: hours_or_note}, 'category': str, 'efir_specific': bool } }
SUBJECTS = {
    # ── Maternelle domains (Cycle 1) ──
    'Mobiliser le langage dans toutes ses dimensions': {
        'levels': {'PS': '~9h', 'MS': '~9h', 'GS': '~9h'},
        'category': 'Domaine d\'apprentissage (Maternelle)',
        'efir_specific': False,
        'notes': 'Inclut oral et écrit. Domaine principal du cycle 1.',
    },
    'Agir, s\'exprimer, comprendre à travers l\'activité physique': {
        'levels': {'PS': '~3h', 'MS': '~3h', 'GS': '~3h'},
        'category': 'Domaine d\'apprentissage (Maternelle)',
        'efir_specific': False,
        'notes': 'Activités motrices, jeux collectifs, expression corporelle.',
    },
    'Agir, s\'exprimer, comprendre à travers les activités artistiques': {
        'levels': {'PS': '~2h', 'MS': '~2h', 'GS': '~2h'},
        'category': 'Domaine d\'apprentissage (Maternelle)',
        'efir_specific': False,
        'notes': 'Arts visuels et éducation musicale.',
    },
    'Acquérir les premiers outils mathématiques': {
        'levels': {'PS': '~3h', 'MS': '~3h', 'GS': '~3h'},
        'category': 'Domaine d\'apprentissage (Maternelle)',
        'efir_specific': False,
        'notes': 'Nombres, formes, grandeurs, suites organisées.',
    },
    'Explorer le monde': {
        'levels': {'PS': '~3h', 'MS': '~3h', 'GS': '~3h'},
        'category': 'Domaine d\'apprentissage (Maternelle)',
        'efir_specific': False,
        'notes': 'Temps, espace, vivant, matière, objets, outils numériques.',
    },

    # ── Core subjects (Élémentaire → Lycée) ──
    'Français': {
        'levels': {
            'CP': '10h', 'CE1': '10h', 'CE2': '10h',
            'CM1': '8h', 'CM2': '8h',
            '6ème': '4h30', '5ème': '4h30', '4ème': '4h30', '3ème': '4h',
            '2nde': '4h', '1ère': '4h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Matière principale. En Terminale, remplacé par Philosophie.',
    },
    'Mathématiques': {
        'levels': {
            'CP': '5h', 'CE1': '5h', 'CE2': '5h',
            'CM1': '5h', 'CM2': '5h',
            '6ème': '4h30', '5ème': '3h30', '4ème': '3h30', '3ème': '3h30',
            '2nde': '4h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'En 1ère/Terminale, devient spécialité ou Maths complémentaires.',
    },
    'Histoire-Géographie': {
        'levels': {
            'CM1': '2h30', 'CM2': '2h30',
            '6ème': '3h', '5ème': '3h', '4ème': '3h', '3ème': '3h30',
            '2nde': '3h', '1ère': '3h', 'Terminale': '3h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Cycle 2: intégré dans "Questionner le monde".',
    },
    'Enseignement Moral et Civique (EMC)': {
        'levels': {
            'CP': '0h30', 'CE1': '0h30', 'CE2': '0h30',
            'CM1': '0h30', 'CM2': '0h30',
            '6ème': '0h30', '5ème': '0h30', '4ème': '0h30', '3ème': '0h30',
            '2nde': '18h/an', '1ère': '18h/an', 'Terminale': '18h/an',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Souvent intégré à l\'Histoire-Géo au collège. 18h annuelles au lycée.',
    },
    'Éducation Physique et Sportive (EPS)': {
        'levels': {
            'CP': '3h', 'CE1': '3h', 'CE2': '3h',
            'CM1': '3h', 'CM2': '3h',
            '6ème': '4h', '5ème': '3h', '4ème': '3h', '3ème': '3h',
            '2nde': '2h', '1ère': '2h', 'Terminale': '2h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Horaire renforcé en 6ème (4h). ORS spécifique pour prof EPS (20h).',
    },

    # ── Languages ──
    'Anglais (LV1 / LVA)': {
        'levels': {
            'CP': '1h30', 'CE1': '1h30', 'CE2': '1h30',
            'CM1': '1h30', 'CM2': '1h30',
            '6ème': '4h', '5ème': '3h', '4ème': '3h', '3ème': '3h',
            '2nde': '—', '1ère': '—', 'Terminale': '—',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'LV1 au primaire/collège. Au lycée, intégré dans le bloc LVA+LVB.',
    },
    'Langues Vivantes A + B (LVA + LVB)': {
        'levels': {
            '2nde': '5h30', '1ère': '4h30', 'Terminale': '4h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Bloc horaire commun LVA (Anglais) + LVB (Arabe ou autre). Non séparable dans l\'emploi du temps officiel.',
    },
    'Arabe (LV2 / LVB)': {
        'levels': {
            'PS': '~2h', 'MS': '~2h', 'GS': '~2h',
            'CP': '~2h', 'CE1': '~2h', 'CE2': '~2h',
            'CM1': '~2h', 'CM2': '~2h',
            '5ème': '2h30', '4ème': '2h30', '3ème': '2h30',
        },
        'category': 'Enseignement obligatoire (EFIR)',
        'efir_specific': True,
        'notes': 'Obligatoire dès la PS à EFIR (exigence pays hôte KSA). LV2 au collège. Au lycée, intégré dans LVA+LVB.',
    },

    # ── Sciences ──
    'Questionner le monde': {
        'levels': {
            'CP': '2h30', 'CE1': '2h30', 'CE2': '2h30',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Cycle 2 uniquement. Regroupe sciences, espace, temps. Devient "Sciences et technologie" au cycle 3.',
    },
    'Sciences et Technologie': {
        'levels': {
            'CM1': '2h', 'CM2': '2h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Cycle 3 élémentaire. En 6ème, les sciences sont séparées (SVT + Physique-Chimie). Technologie supprimée en 6ème depuis 2024.',
    },
    'Sciences de la Vie et de la Terre (SVT)': {
        'levels': {
            '6ème': '~1h30', '5ème': '1h30', '4ème': '1h30', '3ème': '1h30',
            '2nde': '1h30',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'En 6ème regroupé avec Physique-Chimie (3h total). En 1ère/Term, disponible comme spécialité.',
    },
    'Physique-Chimie': {
        'levels': {
            '6ème': '~1h30', '5ème': '1h30', '4ème': '1h30', '3ème': '1h30',
            '2nde': '3h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'En 6ème regroupé avec SVT (3h total). En 1ère/Term, disponible comme spécialité.',
    },
    'Technologie': {
        'levels': {
            '5ème': '1h30', '4ème': '1h30', '3ème': '1h30',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Supprimée en 6ème depuis la réforme 2024. Maintenue en 5ème-3ème.',
    },
    'Sciences Numériques et Technologie (SNT)': {
        'levels': {
            '2nde': '1h30',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Seconde uniquement. Introduction au numérique, données, web, réseaux.',
    },
    'Enseignement scientifique': {
        'levels': {
            '1ère': '2h', 'Terminale': '2h',
        },
        'category': 'Enseignement obligatoire (Tronc commun)',
        'efir_specific': False,
        'notes': 'Tronc commun 1ère et Terminale. Remplace SVT + Physique-Chimie du tronc commun.',
    },

    # ── Arts ──
    'Enseignements artistiques': {
        'levels': {
            'CP': '2h', 'CE1': '2h', 'CE2': '2h',
            'CM1': '2h', 'CM2': '2h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Arts plastiques + Éducation musicale combinés en élémentaire (1h chacun).',
    },
    'Arts Plastiques': {
        'levels': {
            '6ème': '1h', '5ème': '1h', '4ème': '1h', '3ème': '1h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': '1h/semaine au collège. Optionnel au lycée (non proposé à EFIR).',
    },
    'Éducation Musicale': {
        'levels': {
            '6ème': '1h', '5ème': '1h', '4ème': '1h', '3ème': '1h',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': '1h/semaine au collège. Optionnel au lycée (non proposé à EFIR).',
    },

    # ── Lycée: tronc commun supplémentaire ──
    'Sciences Économiques et Sociales (SES)': {
        'levels': {
            '2nde': '1h30',
        },
        'category': 'Enseignement obligatoire',
        'efir_specific': False,
        'notes': 'Obligatoire en 2nde (découverte). Disponible en spécialité 1ère/Term.',
    },
    'Philosophie': {
        'levels': {
            'Terminale': '4h',
        },
        'category': 'Enseignement obligatoire (Tronc commun)',
        'efir_specific': False,
        'notes': 'Terminale uniquement. Remplace le Français du tronc commun.',
    },

    # ── EFIR-specific ──
    'Éducation islamique': {
        'levels': {
            'PS': '~1h', 'MS': '~1h', 'GS': '~1h',
            'CP': '~1h', 'CE1': '~1h', 'CE2': '~1h',
            'CM1': '~1h', 'CM2': '~1h',
            '6ème': '~1h', '5ème': '~1h', '4ème': '~1h', '3ème': '~1h',
        },
        'category': 'Enseignement obligatoire (EFIR)',
        'efir_specific': True,
        'notes': 'Exigence réglementaire du pays hôte (KSA). Non présent dans le curriculum AEFE standard.',
    },
}

# Spécialités Lycée (1ère: 3 × 4h, Terminale: 2 × 6h)
SPECIALITES = {
    'Mathématiques (Spécialité)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Spécialité la plus choisie. Peut être abandonnée en Terminale → Maths complémentaires.',
    },
    'Physique-Chimie (Spécialité)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Spécialité scientifique principale avec SVT.',
    },
    'Sciences de la Vie et de la Terre (Spécialité)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Associée souvent à Physique-Chimie ou Mathématiques.',
    },
    'Sciences Économiques et Sociales (Spécialité)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Voie économique et sociale.',
    },
    'Histoire-Géographie, Géopolitique et Sciences Politiques (HGGSP)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Spécialité littéraire/sciences humaines.',
    },
    'Humanités, Littérature et Philosophie (HLP)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'Spécialité littéraire.',
    },
    'Langues, Littératures et Cultures Étrangères et Régionales (LLCER)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': True,
        'notes': 'LLCER Anglais à EFIR.',
    },
    'Numérique et Sciences Informatiques (NSI)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Non proposé à EFIR. Présent dans BudFin seed data — à vérifier.',
    },
    'Sciences de l\'Ingénieur (SI)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Non proposé à EFIR.',
    },
    'Arts (Spécialité)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Non proposé à EFIR. Inclut arts plastiques, musique, théâtre, cinéma, etc.',
    },
    'Biologie-Écologie': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Réservé aux lycées agricoles. Non applicable à EFIR.',
    },
    'Littérature, Langues et Cultures de l\'Antiquité (LLCA)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Latin/Grec. Non proposé à EFIR.',
    },
    'Éducation Physique, Pratiques et Culture Sportives (EPPCS)': {
        'levels': {'1ère': '4h', 'Terminale': '6h'},
        'offered_efir': False,
        'notes': 'Nouvelle spécialité (2024). Non proposée à EFIR.',
    },
}

# Options lycée
OPTIONS_LYCEE = {
    'Mathématiques complémentaires': {
        'levels': {'Terminale': '3h'},
        'offered_efir': True,
        'notes': 'Pour les élèves ayant abandonné la spécialité Maths en Terminale.',
    },
    'Mathématiques expertes': {
        'levels': {'Terminale': '3h'},
        'offered_efir': True,
        'notes': 'Pour les élèves gardant la spécialité Maths en Terminale.',
    },
    'Latin': {
        'levels': {'5ème': '2h', '4ème': '2h', '3ème': '2h'},
        'offered_efir': False,
        'notes': 'Option collège. Non proposé à EFIR.',
    },
    'LV3 (Espagnol / Allemand)': {
        'levels': {'5ème': '2h30', '4ème': '2h30', '3ème': '2h30'},
        'offered_efir': False,
        'notes': 'Option collège. Non proposé à EFIR actuellement.',
    },
}

# BudFin existing disciplines for gap analysis
BUDFIN_DISCIPLINES = {
    'FRANCAIS': {'name': 'Francais', 'category': 'SUBJECT'},
    'MATHEMATIQUES': {'name': 'Mathematiques', 'category': 'SUBJECT'},
    'HISTOIRE_GEO': {'name': 'Histoire-Geographie', 'category': 'SUBJECT'},
    'ANGLAIS_LV1': {'name': 'Anglais LV1', 'category': 'SUBJECT'},
    'ARABE': {'name': 'Arabe', 'category': 'SUBJECT'},
    'ISLAMIQUE': {'name': 'Islamique', 'category': 'SUBJECT'},
    'EPS': {'name': 'Education Physique et Sportive', 'category': 'SUBJECT'},
    'PHYSIQUE_CHIMIE': {'name': 'Physique-Chimie', 'category': 'SUBJECT'},
    'SVT': {'name': 'Sciences de la Vie et de la Terre', 'category': 'SUBJECT'},
    'TECHNOLOGIE': {'name': 'Technologie', 'category': 'SUBJECT'},
    'ARTS_PLASTIQUES': {'name': 'Arts Plastiques', 'category': 'SUBJECT'},
    'EDUCATION_MUSICALE': {'name': 'Education Musicale', 'category': 'SUBJECT'},
    'PHILOSOPHIE': {'name': 'Philosophie', 'category': 'SUBJECT'},
    'SES': {'name': 'Sciences Economiques et Sociales', 'category': 'SUBJECT'},
    'NSI': {'name': 'Numerique et Sciences Informatiques', 'category': 'SUBJECT'},
    'PRIMARY_HOMEROOM': {'name': 'Enseignement Primaire', 'category': 'ROLE'},
    'ASEM': {'name': 'Agent Specialise Ecoles Maternelles', 'category': 'ROLE'},
    'AUTONOMY': {'name': "Heures d'Autonomie", 'category': 'POOL'},
}


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_overview(wb):
    """Sheet 1: Master matrix — all subjects × all levels."""
    ws = wb.active
    ws.title = 'Vue d\'ensemble'
    ws.sheet_properties.tabColor = '2C3E50'

    # Title
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'EFIR — Couverture des matières par niveau'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:Q2')
    ws['A2'] = 'Référentiel Éducation Nationale / AEFE — Mars 2026'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='7F8C8D')
    ws['A2'].alignment = CENTER

    # Headers
    row = 4
    headers = ['Matière (Nom officiel)', 'Catégorie'] + LEVELS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Cycle sub-headers in row 5
    row = 5
    ws.cell(row=row, column=1, value='')
    ws.cell(row=row, column=2, value='')
    cycle_labels = {
        3: 'Maternelle', 4: '', 5: '',
        6: 'Élémentaire C2', 7: '', 8: '',
        9: 'Élémentaire C3', 10: '',
        11: 'Collège', 12: '', 13: '', 14: '',
        15: 'Lycée', 16: '', 17: '',
    }
    cycle_fills_map = {
        3: MATERNELLE_FILL, 4: MATERNELLE_FILL, 5: MATERNELLE_FILL,
        6: ELEM_FILL, 7: ELEM_FILL, 8: ELEM_FILL,
        9: ELEM_FILL, 10: ELEM_FILL,
        11: COLLEGE_FILL, 12: COLLEGE_FILL, 13: COLLEGE_FILL, 14: COLLEGE_FILL,
        15: LYCEE_FILL, 16: LYCEE_FILL, 17: LYCEE_FILL,
    }
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = cycle_labels.get(col_idx, '')
        cell.font = BOLD_FONT
        cell.fill = cycle_fills_map.get(col_idx, WHITE_FILL)
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Merge cycle labels
    ws.merge_cells('C5:E5')  # Maternelle
    ws.merge_cells('F5:H5')  # Elem C2
    ws.merge_cells('I5:J5')  # Elem C3
    ws.merge_cells('K5:N5')  # Collège
    ws.merge_cells('O5:Q5')  # Lycée

    # Data rows
    row = 6
    for subj_name, info in SUBJECTS.items():
        ws.cell(row=row, column=1, value=subj_name).font = BOLD_FONT if not info['efir_specific'] else Font(name='Calibri', bold=True, size=10, color='E67E22')
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER

        cat_cell = ws.cell(row=row, column=2, value=info['category'])
        cat_cell.font = NORMAL_FONT
        cat_cell.alignment = LEFT_WRAP
        cat_cell.border = THIN_BORDER
        if info['efir_specific']:
            cat_cell.fill = YELLOW_FILL

        for col_idx, level in enumerate(LEVELS, 3):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if level in info['levels']:
                cell.value = info['levels'][level]
                cell.font = BOLD_FONT
                cell.fill = GREEN_FILL
            else:
                cell.value = '—'
                cell.font = Font(name='Calibri', size=10, color='BDBDBD')
                cell.fill = LIGHT_GRAY_FILL

        row += 1

    # Separator
    row += 1
    ws.merge_cells(f'A{row}:Q{row}')
    ws.cell(row=row, column=1, value='SPÉCIALITÉS LYCÉE (1ère: 3 × 4h  |  Terminale: 2 × 6h)').font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = LYCEE_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    row += 1

    for spec_name, info in SPECIALITES.items():
        ws.cell(row=row, column=1, value=spec_name).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER

        status = 'Proposé à EFIR' if info['offered_efir'] else 'Non proposé à EFIR'
        cat_cell = ws.cell(row=row, column=2, value=status)
        cat_cell.font = NORMAL_FONT
        cat_cell.alignment = LEFT_WRAP
        cat_cell.border = THIN_BORDER
        cat_cell.fill = GREEN_FILL if info['offered_efir'] else RED_FILL

        for col_idx, level in enumerate(LEVELS, 3):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if level in info['levels']:
                cell.value = info['levels'][level]
                cell.font = BOLD_FONT
                if info['offered_efir']:
                    cell.fill = GREEN_FILL
                else:
                    cell.fill = RED_FILL
            else:
                cell.value = '—'
                cell.font = Font(name='Calibri', size=10, color='BDBDBD')
                cell.fill = LIGHT_GRAY_FILL

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # Freeze panes
    ws.freeze_panes = 'C6'

    # Legend
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value='Légende').font = SUBTITLE_FONT
    row += 1
    legend_items = [
        ('Vert', GREEN_FILL, 'Matière enseignée à ce niveau'),
        ('Jaune', YELLOW_FILL, 'Spécifique EFIR (exigence pays hôte KSA)'),
        ('Rouge', RED_FILL, 'Spécialité non proposée à EFIR'),
        ('Gris', LIGHT_GRAY_FILL, 'Non applicable à ce niveau'),
    ]
    for label, fill, desc in legend_items:
        ws.cell(row=row, column=1, value=label).fill = fill
        ws.cell(row=row, column=1).font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=desc).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1


def build_detail_sheet(wb, title, tab_color, levels_subset, desc):
    """Build a detailed sheet for a cycle."""
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = tab_color

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'EFIR — {title}'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:F2')
    ws['A2'] = desc
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='7F8C8D')
    ws['A2'].alignment = CENTER

    row = 4
    headers = ['Matière', 'Catégorie', 'Niveaux concernés', 'Horaires', 'EFIR spécifique', 'Notes']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = 5
    for subj_name, info in SUBJECTS.items():
        # Check if this subject appears in any of the levels for this sheet
        relevant_levels = {lv: hr for lv, hr in info['levels'].items() if lv in levels_subset}
        if not relevant_levels:
            continue

        ws.cell(row=row, column=1, value=subj_name).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=info['category']).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = LEFT_WRAP
        ws.cell(row=row, column=2).border = THIN_BORDER

        levels_str = ', '.join(relevant_levels.keys())
        ws.cell(row=row, column=3, value=levels_str).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER

        hours_str = ' / '.join(f'{lv}: {hr}' for lv, hr in relevant_levels.items())
        ws.cell(row=row, column=4, value=hours_str).font = NORMAL_FONT
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).border = THIN_BORDER

        efir_cell = ws.cell(row=row, column=5)
        if info['efir_specific']:
            efir_cell.value = 'Oui'
            efir_cell.font = WARN_FONT
            efir_cell.fill = YELLOW_FILL
        else:
            efir_cell.value = 'Non'
            efir_cell.font = NORMAL_FONT
        efir_cell.alignment = CENTER
        efir_cell.border = THIN_BORDER

        ws.cell(row=row, column=6, value=info.get('notes', '')).font = NORMAL_FONT
        ws.cell(row=row, column=6).alignment = LEFT_WRAP
        ws.cell(row=row, column=6).border = THIN_BORDER

        row += 1

    # Add specialties for Lycée sheets
    if title in ('Lycée', 'Lycée (2nde-Terminale)'):
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value='SPÉCIALITÉS').font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = LYCEE_FILL
        ws.cell(row=row, column=1).alignment = CENTER
        row += 1

        for spec_name, info in SPECIALITES.items():
            relevant = {lv: hr for lv, hr in info['levels'].items() if lv in levels_subset}
            if not relevant:
                continue

            ws.cell(row=row, column=1, value=spec_name).font = BOLD_FONT
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2, value='Spécialité').font = NORMAL_FONT
            ws.cell(row=row, column=2).alignment = LEFT_WRAP
            ws.cell(row=row, column=2).border = THIN_BORDER

            levels_str = ', '.join(relevant.keys())
            ws.cell(row=row, column=3, value=levels_str).font = NORMAL_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=3).border = THIN_BORDER

            hours_str = ' / '.join(f'{lv}: {hr}' for lv, hr in relevant.items())
            ws.cell(row=row, column=4, value=hours_str).font = NORMAL_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=4).border = THIN_BORDER

            efir_cell = ws.cell(row=row, column=5)
            if info['offered_efir']:
                efir_cell.value = 'Oui'
                efir_cell.font = CHECK_FONT
                efir_cell.fill = GREEN_FILL
            else:
                efir_cell.value = 'Non'
                efir_cell.font = CROSS_FONT
                efir_cell.fill = RED_FILL
            efir_cell.alignment = CENTER
            efir_cell.border = THIN_BORDER

            ws.cell(row=row, column=6, value=info.get('notes', '')).font = NORMAL_FONT
            ws.cell(row=row, column=6).alignment = LEFT_WRAP
            ws.cell(row=row, column=6).border = THIN_BORDER

            row += 1

        # Options
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value='OPTIONS').font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).alignment = CENTER
        row += 1

        for opt_name, info in OPTIONS_LYCEE.items():
            relevant = {lv: hr for lv, hr in info['levels'].items() if lv in levels_subset}
            if not relevant:
                continue

            ws.cell(row=row, column=1, value=opt_name).font = BOLD_FONT
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2, value='Option').font = NORMAL_FONT
            ws.cell(row=row, column=2).alignment = LEFT_WRAP
            ws.cell(row=row, column=2).border = THIN_BORDER

            levels_str = ', '.join(relevant.keys())
            ws.cell(row=row, column=3, value=levels_str).font = NORMAL_FONT
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=3).border = THIN_BORDER

            hours_str = ' / '.join(f'{lv}: {hr}' for lv, hr in relevant.items())
            ws.cell(row=row, column=4, value=hours_str).font = NORMAL_FONT
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=4).border = THIN_BORDER

            efir_cell = ws.cell(row=row, column=5)
            if info['offered_efir']:
                efir_cell.value = 'Oui'
                efir_cell.font = CHECK_FONT
                efir_cell.fill = GREEN_FILL
            else:
                efir_cell.value = 'Non'
                efir_cell.font = CROSS_FONT
                efir_cell.fill = RED_FILL
            efir_cell.alignment = CENTER
            efir_cell.border = THIN_BORDER

            ws.cell(row=row, column=6, value=info.get('notes', '')).font = NORMAL_FONT
            ws.cell(row=row, column=6).alignment = LEFT_WRAP
            ws.cell(row=row, column=6).border = THIN_BORDER

            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 55

    ws.freeze_panes = 'A5'


def build_gap_analysis(wb):
    """Sheet: Gap analysis between EFIR curriculum and BudFin disciplines."""
    ws = wb.create_sheet(title='Analyse des écarts')
    ws.sheet_properties.tabColor = 'E74C3C'

    ws.merge_cells('A1:G1')
    ws['A1'] = 'Analyse des écarts — BudFin vs Curriculum EFIR'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:G2')
    ws['A2'] = 'Comparaison des disciplines dans BudFin avec le référentiel officiel'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='7F8C8D')
    ws['A2'].alignment = CENTER

    # ── Section 1: BudFin disciplines matched to official names ──
    row = 4
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value='1. Disciplines BudFin existantes — Correspondance officielle').font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
    row += 1

    headers = ['Code BudFin', 'Nom BudFin', 'Catégorie', 'Nom officiel (Éducation Nationale)', 'Statut', 'Remarque', 'Action recommandée']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1

    # Mapping BudFin code → official name + status
    gap_data = [
        ('FRANCAIS', 'Francais', 'SUBJECT', 'Français', 'OK', 'Nom correct, manque accent', 'Aucune'),
        ('MATHEMATIQUES', 'Mathematiques', 'SUBJECT', 'Mathématiques', 'OK', 'Nom correct, manque accent', 'Aucune'),
        ('HISTOIRE_GEO', 'Histoire-Geographie', 'SUBJECT', 'Histoire-Géographie', 'OK', 'Nom correct, manque accent', 'Aucune'),
        ('ANGLAIS_LV1', 'Anglais LV1', 'SUBJECT', 'Anglais (LV1 / LVA)', 'OK', 'Au lycée, devient LVA dans le bloc LVA+LVB', 'Ajouter alias "LVA Anglais"'),
        ('ARABE', 'Arabe', 'SUBJECT', 'Arabe (LV2 / LVB)', 'OK', 'Spécifique EFIR + KSA. LVB au lycée.', 'Aucune'),
        ('ISLAMIQUE', 'Islamique', 'SUBJECT', 'Éducation islamique', 'RENOMMER', 'Nom officiel: "Éducation islamique"', 'Renommer en "Éducation islamique"'),
        ('EPS', 'Education Physique et Sportive', 'SUBJECT', 'Éducation Physique et Sportive (EPS)', 'OK', 'Nom complet correct', 'Aucune'),
        ('PHYSIQUE_CHIMIE', 'Physique-Chimie', 'SUBJECT', 'Physique-Chimie', 'OK', 'Exact', 'Aucune'),
        ('SVT', 'Sciences de la Vie et de la Terre', 'SUBJECT', 'Sciences de la Vie et de la Terre (SVT)', 'OK', 'Exact', 'Aucune'),
        ('TECHNOLOGIE', 'Technologie', 'SUBJECT', 'Technologie', 'OK', 'Supprimée en 6ème (réforme 2024). Maintenue 5è-3è.', 'Vérifier DHG rules pour 6ème'),
        ('ARTS_PLASTIQUES', 'Arts Plastiques', 'SUBJECT', 'Arts Plastiques', 'OK', 'Exact. Collège uniquement à EFIR.', 'Aucune'),
        ('EDUCATION_MUSICALE', 'Education Musicale', 'SUBJECT', 'Éducation Musicale', 'OK', 'Manque accent', 'Aucune'),
        ('PHILOSOPHIE', 'Philosophie', 'SUBJECT', 'Philosophie', 'OK', 'Terminale uniquement', 'Aucune'),
        ('SES', 'Sciences Economiques et Sociales', 'SUBJECT', 'Sciences Économiques et Sociales (SES)', 'OK', 'Manque accent', 'Aucune'),
        ('NSI', 'Numerique et Sciences Informatiques', 'SUBJECT', 'Numérique et Sciences Informatiques (NSI)', 'VÉRIFIER', 'Présent dans BudFin mais NON proposé à EFIR', 'Confirmer si EFIR propose NSI ou supprimer'),
        ('PRIMARY_HOMEROOM', 'Enseignement Primaire', 'ROLE', 'Professeur des Écoles (PE)', 'OK', 'Rôle, pas une matière', 'Aucune'),
        ('ASEM', 'Agent Specialise Ecoles Maternelles', 'ROLE', 'Agent Spécialisé des Écoles Maternelles (ATSEM)', 'OK', 'Rôle, pas une matière. Nom officiel: ATSEM.', 'Ajouter alias "ATSEM"'),
        ('AUTONOMY', "Heures d'Autonomie", 'POOL', "Heures d'Autonomie", 'OK', 'Pool horaire, pas une matière', 'Aucune'),
    ]

    for code, name, cat, official, status, remark, action in gap_data:
        ws.cell(row=row, column=1, value=code).font = Font(name='Calibri', bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=cat).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=official).font = BOLD_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER

        status_cell = ws.cell(row=row, column=5, value=status)
        status_cell.alignment = CENTER
        status_cell.border = THIN_BORDER
        if status == 'OK':
            status_cell.font = CHECK_FONT
            status_cell.fill = GREEN_FILL
        elif status == 'RENOMMER':
            status_cell.font = WARN_FONT
            status_cell.fill = YELLOW_FILL
        else:
            status_cell.font = CROSS_FONT
            status_cell.fill = RED_FILL

        ws.cell(row=row, column=6, value=remark).font = NORMAL_FONT
        ws.cell(row=row, column=6).alignment = LEFT_WRAP
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=7, value=action).font = NORMAL_FONT
        ws.cell(row=row, column=7).alignment = LEFT_WRAP
        ws.cell(row=row, column=7).border = THIN_BORDER

        row += 1

    # ── Section 2: Missing from BudFin ──
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value='2. Matières enseignées à EFIR mais ABSENTES de BudFin').font = SUBTITLE_FONT
    ws.cell(row=row, column=1).fill = RED_FILL
    row += 1

    headers2 = ['Matière', 'Niveaux', 'Horaires', 'Type', 'Priorité', 'Code suggéré', 'Notes']
    for col_idx, h in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1

    missing = [
        ('Enseignement Moral et Civique (EMC)', 'CP → Terminale', '0h30/sem (primaire/collège), 18h/an (lycée)', 'Tronc commun', 'Moyenne', 'EMC', 'Souvent mutualisé avec Histoire-Géo. Peut rester intégré.'),
        ('Sciences Numériques et Technologie (SNT)', '2nde', '1h30/sem', 'Tronc commun', 'Haute', 'SNT', 'Obligatoire en 2nde. Nécessite un enseignant dédié ou mutualisé.'),
        ('Enseignement scientifique', '1ère, Terminale', '2h/sem', 'Tronc commun', 'Haute', 'ENS_SCIENTIFIQUE', 'Tronc commun obligatoire. Enseigné par profs SVT/Physique.'),
        ('Questionner le monde', 'CP, CE1, CE2', '2h30/sem', 'Élémentaire C2', 'Basse', 'QUESTIONNER_MONDE', 'Cycle 2. Enseigné par PE (PRIMARY_HOMEROOM). Pas de discipline séparée nécessaire.'),
        ('Sciences et Technologie', 'CM1, CM2', '2h/sem', 'Élémentaire C3', 'Basse', 'SCIENCES_TECHNO', 'Cycle 3 élémentaire. Enseigné par PE. Pas de discipline séparée nécessaire.'),
        ('Enseignements artistiques (Élémentaire)', 'CP → CM2', '2h/sem (1h AP + 1h EM)', 'Élémentaire', 'Basse', '—', 'Enseigné par PE. AP et EM existent déjà en tant que disciplines collège.'),
        ('HGGSP (Spécialité)', '1ère, Terminale', '4h / 6h', 'Spécialité', 'Haute', 'HGGSP', 'Proposé à EFIR. Enseigné par prof Histoire-Géo.'),
        ('HLP (Spécialité)', '1ère, Terminale', '4h / 6h', 'Spécialité', 'Haute', 'HLP', 'Proposé à EFIR. Enseigné par prof Français/Philosophie.'),
        ('LLCER (Spécialité)', '1ère, Terminale', '4h / 6h', 'Spécialité', 'Haute', 'LLCER', 'Proposé à EFIR. LLCER Anglais.'),
        ('Mathématiques complémentaires', 'Terminale', '3h', 'Option', 'Moyenne', 'MATHS_COMP', 'Pour élèves ayant abandonné spé Maths.'),
        ('Mathématiques expertes', 'Terminale', '3h', 'Option', 'Moyenne', 'MATHS_EXPERTES', 'Pour élèves gardant spé Maths.'),
    ]

    for name, niveaux, hours, type_, priority, code, notes in missing:
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = LEFT_WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=niveaux).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER

        ws.cell(row=row, column=3, value=hours).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=3).border = THIN_BORDER

        ws.cell(row=row, column=4, value=type_).font = NORMAL_FONT
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).border = THIN_BORDER

        prio_cell = ws.cell(row=row, column=5, value=priority)
        prio_cell.alignment = CENTER
        prio_cell.border = THIN_BORDER
        if priority == 'Haute':
            prio_cell.font = CROSS_FONT
            prio_cell.fill = RED_FILL
        elif priority == 'Moyenne':
            prio_cell.font = WARN_FONT
            prio_cell.fill = YELLOW_FILL
        else:
            prio_cell.font = NORMAL_FONT
            prio_cell.fill = GREEN_FILL

        ws.cell(row=row, column=6, value=code).font = Font(name='Courier New', size=10)
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=6).border = THIN_BORDER

        ws.cell(row=row, column=7, value=notes).font = NORMAL_FONT
        ws.cell(row=row, column=7).alignment = LEFT_WRAP
        ws.cell(row=row, column=7).border = THIN_BORDER

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 50

    ws.freeze_panes = 'A6'


def build_budfin_current(wb):
    """Sheet: Current BudFin discipline seed data."""
    ws = wb.create_sheet(title='BudFin (données actuelles)')
    ws.sheet_properties.tabColor = '27AE60'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'BudFin — Disciplines actuelles (seed data)'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:F2')
    ws['A2'] = 'Source: apps/api/prisma/seeds/staffing-master-data.ts'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='7F8C8D')
    ws['A2'].alignment = CENTER

    row = 4
    headers = ['Code', 'Nom', 'Catégorie', 'Sort Order', 'Alias existants', 'Nom officiel (Éducation Nationale)']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    aliases = {
        'MATHEMATIQUES': 'Mathematiques, Maths',
        'HISTOIRE_GEO': 'Histoire-Geographie, Histoire Géographie',
        'EPS': 'Education Physique',
        'PHYSIQUE_CHIMIE': 'Sciences Physiques',
        'ARTS_PLASTIQUES': 'Arts Visuels',
    }

    official_names = {
        'FRANCAIS': 'Français',
        'MATHEMATIQUES': 'Mathématiques',
        'HISTOIRE_GEO': 'Histoire-Géographie',
        'ANGLAIS_LV1': 'Anglais (LV1 / LVA)',
        'ARABE': 'Arabe (LV2 / LVB)',
        'ISLAMIQUE': 'Éducation islamique',
        'EPS': 'Éducation Physique et Sportive (EPS)',
        'PHYSIQUE_CHIMIE': 'Physique-Chimie',
        'SVT': 'Sciences de la Vie et de la Terre (SVT)',
        'TECHNOLOGIE': 'Technologie',
        'ARTS_PLASTIQUES': 'Arts Plastiques',
        'EDUCATION_MUSICALE': 'Éducation Musicale',
        'PHILOSOPHIE': 'Philosophie',
        'SES': 'Sciences Économiques et Sociales (SES)',
        'NSI': 'Numérique et Sciences Informatiques (NSI)',
        'PRIMARY_HOMEROOM': 'Professeur des Écoles (PE)',
        'ASEM': 'Agent Spécialisé des Écoles Maternelles (ATSEM)',
        'AUTONOMY': "Heures d'Autonomie",
    }

    row = 5
    sort_order = 1
    for code, info in BUDFIN_DISCIPLINES.items():
        ws.cell(row=row, column=1, value=code).font = Font(name='Courier New', bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=info['name']).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        cat_cell = ws.cell(row=row, column=3, value=info['category'])
        cat_cell.font = NORMAL_FONT
        cat_cell.alignment = CENTER
        cat_cell.border = THIN_BORDER
        if info['category'] == 'SUBJECT':
            cat_cell.fill = LIGHT_BLUE_FILL
        elif info['category'] == 'ROLE':
            cat_cell.fill = YELLOW_FILL
        else:
            cat_cell.fill = LIGHT_GRAY_FILL

        ws.cell(row=row, column=4, value=sort_order).font = NORMAL_FONT
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).border = THIN_BORDER

        ws.cell(row=row, column=5, value=aliases.get(code, '—')).font = NORMAL_FONT
        ws.cell(row=row, column=5).alignment = LEFT_WRAP
        ws.cell(row=row, column=5).border = THIN_BORDER

        ws.cell(row=row, column=6, value=official_names.get(code, '—')).font = NORMAL_FONT
        ws.cell(row=row, column=6).alignment = LEFT_WRAP
        ws.cell(row=row, column=6).border = THIN_BORDER

        sort_order += 1
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 45

    ws.freeze_panes = 'A5'


def build_ors_profiles(wb):
    """Sheet: Service obligation profiles (ORS)."""
    ws = wb.create_sheet(title='Profils ORS')
    ws.sheet_properties.tabColor = '8E44AD'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Profils d\'Obligation Réglementaire de Service (ORS)'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:F2')
    ws['A2'] = 'Heures de service hebdomadaire par type d\'enseignant'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='7F8C8D')
    ws['A2'].alignment = CENTER

    row = 4
    headers = ['Code', 'Nom officiel', 'ORS (h/sem)', 'HSA éligible', 'Niveaux d\'intervention', 'Notes']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = 5
    profiles = [
        ('PE', 'Professeur des Écoles', 24, 'Non', 'Maternelle, Élémentaire',
         '24h devant élèves + 108h annualisées (APC, réunions, formation).'),
        ('CERTIFIÉ', 'Professeur Certifié', 18, 'Oui', 'Collège, Lycée',
         '18h devant élèves. HSA possibles au-delà. Concours: CAPES/CAPET.'),
        ('AGRÉGÉ', 'Professeur Agrégé', 15, 'Oui', 'Collège, Lycée (priorité)',
         '15h devant élèves. HSA possibles. Concours: Agrégation.'),
        ('EPS', 'Professeur d\'EPS', 20, 'Oui', 'Collège, Lycée',
         '20h devant élèves. ORS spécifique EPS. Concours: CAPEPS.'),
        ('ARABE/ISLAMIQUE', 'Enseignant Arabe / Islamique', 24, 'Non', 'Tous niveaux',
         'Recrutement local (KSA). ORS aligné sur PE. Non éligible HSA.'),
        ('ASEM', 'Agent Spécialisé des Écoles Maternelles (ATSEM)', 0, 'Non', 'Maternelle',
         'Personnel non enseignant. Accompagne les PE en maternelle.'),
        ('DOCUMENTALISTE', 'Professeur Documentaliste', 30, 'Non', 'Collège, Lycée',
         '30h hebdomadaires au CDI. Pas d\'enseignement direct (sauf séquences ponctuelles).'),
    ]

    for code, name, ors, hsa, niveaux, notes in profiles:
        ws.cell(row=row, column=1, value=code).font = Font(name='Courier New', bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=name).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        ors_cell = ws.cell(row=row, column=3, value=f'{ors}h')
        ors_cell.font = Font(name='Calibri', bold=True, size=11, color='2C3E50')
        ors_cell.alignment = CENTER
        ors_cell.border = THIN_BORDER

        hsa_cell = ws.cell(row=row, column=4, value=hsa)
        hsa_cell.alignment = CENTER
        hsa_cell.border = THIN_BORDER
        if hsa == 'Oui':
            hsa_cell.font = CHECK_FONT
            hsa_cell.fill = GREEN_FILL
        else:
            hsa_cell.font = CROSS_FONT
            hsa_cell.fill = RED_FILL

        ws.cell(row=row, column=5, value=niveaux).font = NORMAL_FONT
        ws.cell(row=row, column=5).alignment = LEFT_WRAP
        ws.cell(row=row, column=5).border = THIN_BORDER

        ws.cell(row=row, column=6, value=notes).font = NORMAL_FONT
        ws.cell(row=row, column=6).alignment = LEFT_WRAP
        ws.cell(row=row, column=6).border = THIN_BORDER

        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 60

    ws.freeze_panes = 'A5'


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    # Sheet 1: Master overview matrix
    build_overview(wb)

    # Sheet 2-5: Cycle details
    build_detail_sheet(wb, 'Maternelle (PS-GS)', '27AE60',
                       ['PS', 'MS', 'GS'],
                       'Cycle 1 — Apprentissages premiers (24h/semaine)')
    build_detail_sheet(wb, 'Élémentaire (CP-CM2)', '2980B9',
                       ['CP', 'CE1', 'CE2', 'CM1', 'CM2'],
                       'Cycles 2 et 3 — Apprentissages fondamentaux et consolidation (24h/semaine)')
    build_detail_sheet(wb, 'Collège (6è-3è)', 'E67E22',
                       ['6ème', '5ème', '4ème', '3ème'],
                       'Cycles 3 (6ème) et 4 — Approfondissements (25-26h/semaine)')
    build_detail_sheet(wb, 'Lycée (2nde-Terminale)', '8E44AD',
                       ['2nde', '1ère', 'Terminale'],
                       'Détermination + Cycle terminal — Spécialisation (26-28h/semaine)')

    # Sheet 6: Gap analysis
    build_gap_analysis(wb)

    # Sheet 7: BudFin current data
    build_budfin_current(wb)

    # Sheet 8: ORS profiles
    build_ors_profiles(wb)

    # Save
    output_path = '/Users/fakerhelali/Desktop/budfin/data/Subjects/EFIR-Subjects-Coverage.xlsx'
    wb.save(output_path)
    print(f'Excel file saved to: {output_path}')


if __name__ == '__main__':
    main()
